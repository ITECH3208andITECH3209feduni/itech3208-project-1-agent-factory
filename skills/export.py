# skills/export.py
# ──────────────────────────────────────────────────────────────
# Export Skill — PROJ-191
# Exports any SkillResult data to PDF or Excel/CSV format.
#
# PDF:   uses fpdf2 (pure Python); fallback → plain text file
# Excel: uses openpyxl; fallback → CSV
#
# Standalone helpers (importable):
#   export_to_pdf(data, filename=None)   → file path
#   export_to_excel(data, filename=None) → file path
# ──────────────────────────────────────────────────────────────

from __future__ import annotations

import csv
import json
import os
import re
from datetime import datetime
from typing import Any

from skills.base_skill import BaseSkill, SkillResult

# ── Optional heavy deps ────────────────────────────────────────
try:
    from fpdf import FPDF          # fpdf2
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ── Exports directory ──────────────────────────────────────────
EXPORTS_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "exports",
)


def _ensure_exports_dir() -> None:
    os.makedirs(EXPORTS_DIR, exist_ok=True)


def _ts() -> str:
    """Return a filesystem-safe timestamp string."""
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _safe_name(name: str) -> str:
    """Strip characters that are not safe in file names."""
    return re.sub(r"[^\w\-]", "_", name)[:40]


def _pdf_safe(text: str) -> str:
    """Replace Unicode characters unsupported by Helvetica with ASCII equivalents."""
    replacements = [
        ("—", "--"),   # em dash
        ("–", "-"),    # en dash
        ("’", "'"),    # right single quotation mark
        ("‘", "'"),    # left single quotation mark
        ("“", '"'),    # left double quotation mark
        ("”", '"'),    # right double quotation mark
        ("•", "*"),    # bullet
        ("…", "..."),  # horizontal ellipsis
        (" ", " "),    # non-breaking space
    ]
    s = str(text)
    for src, dst in replacements:
        s = s.replace(src, dst)
    return s.encode("latin-1", errors="replace").decode("latin-1")


# ══════════════════════════════════════════════════════════════
# PDF EXPORT
# ══════════════════════════════════════════════════════════════

def export_to_pdf(data: dict, filename: str | None = None) -> str:
    """
    Export a SkillResult dict to PDF.

    Args:
        data:     A dict produced by SkillResult.to_dict().
        filename: Optional base filename (without extension).

    Returns:
        Absolute path of the written file.
    """
    _ensure_exports_dir()

    skill    = data.get("skill", "export")
    ts_str   = _ts()
    base     = filename or f"{_safe_name(skill)}_{ts_str}"
    out_path = os.path.join(EXPORTS_DIR, f"{base}.pdf")

    if FPDF_AVAILABLE:
        _write_fpdf(data, out_path)
    else:
        # Graceful degradation: plain text
        txt_path = out_path.replace(".pdf", ".txt")
        _write_plain_text(data, txt_path)
        return txt_path

    return out_path


def _write_fpdf(data: dict, path: str) -> None:
    """Render a SkillResult dict to a PDF file using fpdf2."""

    class AgentPDF(FPDF):
        def header(self):
            self.set_font("Helvetica", "B", 12)
            self.set_fill_color(60, 80, 140)
            self.set_text_color(255, 255, 255)
            self.cell(0, 10, "Agent Factory -- Export Report", fill=True,
                      new_x="LMARGIN", new_y="NEXT")
            self.set_text_color(0, 0, 0)
            self.ln(2)

        def footer(self):
            self.set_y(-15)
            self.set_font("Helvetica", "I", 8)
            self.cell(0, 10, f"Page {self.page_no()}", align="C")

    pdf = AgentPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    skill    = _pdf_safe(data.get("skill", "unknown"))
    query    = _pdf_safe(data.get("query", ""))
    success  = data.get("success", False)
    summary  = _pdf_safe(data.get("summary", ""))
    error    = _pdf_safe(data.get("error", ""))
    duration = _pdf_safe(str(data.get("duration", "")))
    results  = data.get("results", [])
    metadata = data.get("metadata", {})

    # ── Title & meta ───────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, f"Skill: {skill.upper()}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 6, f"Query   : {query}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Status  : {'SUCCESS' if success else 'FAILED'}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Duration: {duration}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    # ── Error (if any) ─────────────────────────────────────────
    if error:
        pdf.set_fill_color(255, 220, 220)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 8, "Error", fill=True, new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        pdf.multi_cell(0, 6, error)
        pdf.ln(2)

    # ── Summary section ────────────────────────────────────────
    if summary:
        pdf.set_fill_color(230, 240, 255)
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 8, "AI Summary", fill=True, new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        pdf.multi_cell(0, 6, summary)
        pdf.ln(4)

    # ── Results table ──────────────────────────────────────────
    if results:
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 8, f"Results ({len(results)} items)", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        # Collect all column names from the result dicts
        all_keys: list[str] = []
        for r in results:
            for k in r.keys():
                if k not in all_keys:
                    all_keys.append(k)

        # Prioritised column order
        priority = ["title", "name", "price", "rating", "authors", "year",
                    "citations", "source", "url", "abstract", "summary"]
        ordered_keys = [k for k in priority if k in all_keys]
        ordered_keys += [k for k in all_keys if k not in ordered_keys]

        # Cap to avoid overcrowding
        display_keys = ordered_keys[:8]

        # Header row
        col_w = (pdf.w - pdf.l_margin - pdf.r_margin) / len(display_keys)
        pdf.set_fill_color(60, 80, 140)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 8)
        for k in display_keys:
            pdf.cell(col_w, 7, _pdf_safe(str(k).title()[:18]), border=1, fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)

        # Data rows
        pdf.set_font("Helvetica", "", 8)
        alt = False
        for row in results[:50]:
            pdf.set_fill_color(240, 245, 255) if alt else pdf.set_fill_color(255, 255, 255)
            for k in display_keys:
                val = _pdf_safe(str(row.get(k, "")))[:30]
                pdf.cell(col_w, 6, val, border=1, fill=True)
            pdf.ln()
            alt = not alt

        if len(results) > 50:
            pdf.set_font("Helvetica", "I", 9)
            pdf.cell(
                0, 6,
                f"... and {len(results) - 50} more rows (truncated in PDF, use Excel for full data).",
                new_x="LMARGIN", new_y="NEXT",
            )
        pdf.ln(4)

    # ── Metadata section ───────────────────────────────────────
    if metadata:
        pdf.set_fill_color(245, 245, 245)
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 8, "Metadata", fill=True, new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 9)
        for k, v in metadata.items():
            line = _pdf_safe(f"{k}: {v}")[:120]
            pdf.cell(0, 6, line, new_x="LMARGIN", new_y="NEXT")

    pdf.output(path)


def _write_plain_text(data: dict, path: str) -> None:
    """Fallback when fpdf2 is not available: write a plain text report."""
    lines = [
        "=" * 60,
        "Agent Factory -- Export Report",
        "=" * 60,
        f"Skill    : {data.get('skill', '')}",
        f"Query    : {data.get('query', '')}",
        f"Success  : {data.get('success', '')}",
        f"Duration : {data.get('duration', '')}",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
    ]
    if data.get("summary"):
        lines += ["--- Summary ---", data["summary"], ""]
    if data.get("error"):
        lines += ["--- Error ---", data["error"], ""]
    results = data.get("results", [])
    if results:
        lines.append(f"--- Results ({len(results)}) ---")
        for i, r in enumerate(results, 1):
            lines.append(f"\n[{i}]")
            for k, v in r.items():
                lines.append(f"  {k}: {v}")
    if data.get("metadata"):
        lines += ["", "--- Metadata ---"]
        for k, v in data["metadata"].items():
            lines.append(f"  {k}: {v}")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ══════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════

def export_to_excel(data: dict, filename: str | None = None) -> str:
    """
    Export a SkillResult dict to Excel (.xlsx) or CSV fallback.

    Args:
        data:     A dict produced by SkillResult.to_dict().
        filename: Optional base filename (without extension).

    Returns:
        Absolute path of the written file.
    """
    _ensure_exports_dir()

    skill    = data.get("skill", "export")
    ts_str   = _ts()
    base     = filename or f"{_safe_name(skill)}_{ts_str}"

    if OPENPYXL_AVAILABLE:
        out_path = os.path.join(EXPORTS_DIR, f"{base}.xlsx")
        _write_openpyxl(data, out_path)
    else:
        out_path = os.path.join(EXPORTS_DIR, f"{base}.csv")
        _write_csv(data, out_path)

    return out_path


def _write_openpyxl(data: dict, path: str) -> None:
    """Write SkillResult data to a multi-sheet Excel workbook."""

    results  = data.get("results", [])
    metadata = data.get("metadata", {})

    wb = openpyxl.Workbook()

    # ── Sheet 1: Metadata ──────────────────────────────────────
    ws_meta = wb.active
    ws_meta.title = "Metadata"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="3C508C")

    meta_rows = [
        ("Field", "Value"),
        ("Skill", data.get("skill", "")),
        ("Query", data.get("query", "")),
        ("Success", str(data.get("success", ""))),
        ("Duration", data.get("duration", "")),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Results", len(results)),
        ("Summary", data.get("summary", "")[:500]),
        ("Error", data.get("error", "")),
    ]
    for k, v in metadata.items():
        meta_rows.append((str(k), str(v)))

    for row_idx, (k, v) in enumerate(meta_rows, 1):
        c1 = ws_meta.cell(row=row_idx, column=1, value=k)
        c2 = ws_meta.cell(row=row_idx, column=2, value=v)
        if row_idx == 1:
            c1.font = header_font
            c2.font = header_font
            c1.fill = header_fill
            c2.fill = header_fill

    ws_meta.column_dimensions["A"].width = 20
    ws_meta.column_dimensions["B"].width = 60

    # ── Sheet 2: Results ───────────────────────────────────────
    ws_results = wb.create_sheet("Results")

    if results:
        all_keys: list[str] = []
        for r in results:
            for k in r.keys():
                if k not in all_keys:
                    all_keys.append(k)

        # Header
        for col_idx, k in enumerate(all_keys, 1):
            cell = ws_results.cell(row=1, column=col_idx, value=str(k).title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data
        alt_fill = PatternFill("solid", fgColor="EDF2FF")
        for row_idx, row in enumerate(results, 2):
            for col_idx, k in enumerate(all_keys, 1):
                val = row.get(k, "")
                if isinstance(val, (list, dict)):
                    val = json.dumps(val, ensure_ascii=False)
                cell = ws_results.cell(row=row_idx, column=col_idx, value=val)
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # Auto-fit column widths (capped at 60)
        for col_idx, k in enumerate(all_keys, 1):
            max_len = max(
                len(str(k)),
                max((len(str(results[r].get(k, ""))) for r in range(min(len(results), 20))), default=0),
            )
            ws_results.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)
    else:
        ws_results.cell(row=1, column=1, value="No results.")

    # ── Sheet 3: Summary ───────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    summary_text = data.get("summary", "No summary available.")
    # Split into lines for readability
    for line_idx, line in enumerate(summary_text.split("\n"), 1):
        ws_summary.cell(row=line_idx, column=1, value=line)
    ws_summary.column_dimensions["A"].width = 100

    wb.save(path)


def _write_csv(data: dict, path: str) -> None:
    """Fallback when openpyxl is not available: write results to CSV."""
    results = data.get("results", [])

    if not results:
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["field", "value"])
            for k, v in data.items():
                if k != "results":
                    writer.writerow([k, str(v)])
        return

    all_keys: list[str] = []
    for r in results:
        for k in r.keys():
            if k not in all_keys:
                all_keys.append(k)

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=all_keys, extrasaction="ignore")
        writer.writeheader()
        for row in results:
            safe_row = {
                k: (json.dumps(v) if isinstance(v, (list, dict)) else v)
                for k, v in row.items()
            }
            writer.writerow(safe_row)


# ══════════════════════════════════════════════════════════════
# SKILL CLASS
# ══════════════════════════════════════════════════════════════

class ExportSkill(BaseSkill):
    name        = "export"
    description = (
        "Export skill results to PDF or Excel/CSV. "
        "Use when the user wants to download, save, or generate a report of results. "
        "Supports PDF (formatted report) and Excel/CSV (structured data)."
    )
    triggers = [
        "export", "download", "save", "pdf", "excel", "spreadsheet",
        "report", "generate report", "save results", "export results",
        "download results", "save to file", "export to",
    ]

    _INSTRUCTIONS = (
        "To export results, send a message in one of these formats:\n"
        "  * export pdf: <paste JSON data here>\n"
        "  * export excel: <paste JSON data here>\n\n"
        "The JSON data should be a SkillResult dict "
        "(from a previous skill response).\n"
        "Exported files are saved to the exports/ directory."
    )

    def run(self, query: str) -> SkillResult:
        q = query.strip()

        # ── Detect format ──────────────────────────────────────
        fmt      = None
        json_str = ""

        lower = q.lower()
        if lower.startswith("export pdf:") or lower.startswith("pdf:"):
            fmt = "pdf"
            json_str = re.split(r"pdf\s*:", q, maxsplit=1, flags=re.IGNORECASE)[-1].strip()
        elif lower.startswith("export excel:") or lower.startswith("excel:") or \
             lower.startswith("export spreadsheet:") or lower.startswith("spreadsheet:"):
            fmt = "excel"
            json_str = re.split(r"(?:excel|spreadsheet)\s*:", q, maxsplit=1, flags=re.IGNORECASE)[-1].strip()
        elif "pdf" in lower:
            fmt = "pdf"
            # Try to extract a JSON blob anywhere in the query
            json_str = _extract_json(q)
        elif any(w in lower for w in ("excel", "spreadsheet", "xlsx")):
            fmt = "excel"
            json_str = _extract_json(q)
        else:
            # No format specified — return instructions
            return SkillResult(
                skill_name=self.name,
                query=query,
                success=True,
                summary=self._INSTRUCTIONS,
                metadata={"mode": "instructions"},
            )

        # ── Parse JSON data ────────────────────────────────────
        if not json_str:
            return SkillResult(
                skill_name=self.name,
                query=query,
                success=False,
                error="No JSON data found in the query. " + self._INSTRUCTIONS,
            )

        try:
            data = json.loads(json_str)
        except json.JSONDecodeError as exc:
            return SkillResult(
                skill_name=self.name,
                query=query,
                success=False,
                error=f"Could not parse JSON: {exc}. Make sure you paste valid JSON after the colon.",
            )

        # Ensure data is a dict
        if not isinstance(data, dict):
            data = {"results": data if isinstance(data, list) else [data],
                    "skill": "unknown", "query": query, "success": True}

        # ── Export ─────────────────────────────────────────────
        try:
            if fmt == "pdf":
                path = export_to_pdf(data)
                lib  = "fpdf2" if FPDF_AVAILABLE else "plain text (fpdf2 not installed)"
                return SkillResult(
                    skill_name=self.name,
                    query=query,
                    success=True,
                    summary=f"PDF exported successfully using {lib}.\nFile: {path}",
                    metadata={"format": "pdf", "path": path, "library": lib},
                )
            else:  # excel
                path = export_to_excel(data)
                lib  = "openpyxl" if OPENPYXL_AVAILABLE else "CSV (openpyxl not installed)"
                return SkillResult(
                    skill_name=self.name,
                    query=query,
                    success=True,
                    summary=f"Excel exported successfully using {lib}.\nFile: {path}",
                    metadata={"format": "excel", "path": path, "library": lib},
                )
        except Exception as exc:
            return SkillResult(
                skill_name=self.name,
                query=query,
                success=False,
                error=f"Export failed: {exc}",
            )


# ── Helper ─────────────────────────────────────────────────────

def _extract_json(text: str) -> str:
    """Try to find and extract a JSON object or array from a free-form string."""
    # Look for { ... } or [ ... ] patterns
    for pattern in (r"\{.*\}", r"\[.*\]"):
        match = re.search(pattern, text, re.DOTALL)
        if match:
            return match.group()
    return ""
